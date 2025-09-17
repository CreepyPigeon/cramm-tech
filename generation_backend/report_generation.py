import pandas as pd
from openpyxl import load_workbook
import uuid

def generate_uuid():
    return str(uuid.uuid4())

def read_df():
    return pd.read_csv('WS_DER_OTC_TOV_csv_col.csv')

def extract_fx_data(df, instrument="Spot", currency_pair=("USD", "AUD")):
    datesfrom = {}

    # Unpack currency pair
    curr1, curr2 = currency_pair

    # Define base filter
    base_filter = (
        (df["Risk category"] == "Foreign exchange") &
        (df["Instrument"] == instrument) &
        (df["Reporting country"] == "All countries (total)") &
        (df["DER_CURR_LEG1"] == curr1) &
        (df["DER_CURR_LEG2"] == curr2)
    )

    # Updated excluded sectors list
    excluded_sectors = [
        "Other financial institutions",
        "Reporting dealers",
        "Non-reporting banks",
        "Institutional investors",
        "Hedge funds and proprietary trading firms",
        "Official sector financial institutions",
        "Undistributed"
    ]

    # Execution methods for broke_non_bank_ele
    non_bank_ele_exec_methods = [
        "Electronic - direct - other",
        "Electronic - indirect - Reuters EBS",
        "Technical residual (Electronic - indirect - Disclosed venues)",
        "Electronic - indirect - other",
        "Electronic - indirect - Dark pools",
        "Electronic - indirect - other ECN",
        "Electronic - indirect - Disclosed venues"
    ]

    # Define all filtered dataframes
    datesfrom["total"] = df[base_filter]

    # Reporting dealers
    datesfrom["reporting_dealers"] = df[base_filter & (df["Counterparty sector"] == "Reporting dealers")]
    datesfrom["reporting_dealers_local"] = datesfrom["reporting_dealers"][
        datesfrom["reporting_dealers"]["Counterparty country"] == "Residents/Local"
    ]
    datesfrom["reporting_dealers_cross_border"] = datesfrom["reporting_dealers"][
        datesfrom["reporting_dealers"]["Counterparty country"] == "Non-residents/Cross-border"
    ]

    # Other financial institutions
    datesfrom["other_fin"] = df[base_filter & (df["Counterparty sector"] == "Other financial institutions")]
    datesfrom["other_fin_local"] = datesfrom["other_fin"][
        datesfrom["other_fin"]["Counterparty country"] == "Residents/Local"
    ]
    datesfrom["other_fin_cross_border"] = datesfrom["other_fin"][
        datesfrom["other_fin"]["Counterparty country"] == "Non-residents/Cross-border"
    ]

    # Non-reporting banks, institutional investors, hedge funds, official sector, others
    datesfrom["other_fin_non_reporting"] = df[base_filter & (df["Counterparty sector"] == "Non-reporting banks")]
    datesfrom["other_fin_inst_investor"] = df[base_filter & (df["Counterparty sector"] == "Institutional investors")]
    datesfrom["other_fin_hedge"] = df[base_filter & (df["Counterparty sector"] == "Hedge funds and proprietary trading firms")]
    datesfrom["other_fin_official"] = df[base_filter & (df["Counterparty sector"] == "Official sector financial institutions")]

    # "Others" category (not in excluded sectors)
    datesfrom["other_fin_others"] = df[base_filter & (~df["Counterparty sector"].isin(excluded_sectors))]

    datesfrom["other_fin_undistributed"] = df[base_filter & (df["Counterparty sector"] == "Undistributed")]

    # Non-financial customers
    datesfrom["non_fin"] = df[base_filter & (df["Counterparty sector"] == "Non-financial customers")]
    datesfrom["non_fin_local"] = datesfrom["non_fin"][
        datesfrom["non_fin"]["Counterparty country"] == "Residents/Local"
    ]
    datesfrom["non_fin_cross_border"] = datesfrom["non_fin"][
        datesfrom["non_fin"]["Counterparty country"] == "Non-residents/Cross-border"
    ]

    # Brokers / Retail
    datesfrom["broke_non_bank_ele"] = df[
    base_filter &
    df["Execution method"].isin(non_bank_ele_exec_methods) &
    (df['Counterparty sector'] == 'Prime brokered')
]

    datesfrom["broke_other"] = df[
        base_filter &
        ~df["Execution method"].isin(non_bank_ele_exec_methods) &
        (df['Counterparty sector'] == 'Prime brokered')
    ]
    datesfrom["broke_retail"] = df[base_filter & (df["Counterparty sector"] == "Retail-driven")]

    #Maturity
    datesfrom["maturity_one_day"] = df[base_filter & (df["Maturity"] == "One day")]

    datesfrom["maturity_one_up_to_seven"] = df[base_filter & (
        (df["Maturity"] == "7 days or less")
    )]

    datesfrom["maturity_seven_up_to_month"] = df[base_filter & (
        df["Maturity"] == "Over 7 days and up to 1 month"
    )]

    datesfrom["maturity_month_up_to_three"] = df[base_filter & (
        df["Maturity"] == "Over 1 month and up to 3 months"
    )]

    datesfrom["maturity_three_month_up_to_six"] = df[base_filter & (
        df["Maturity"] == "Over 3 months and up to 6 months"
    )]

    datesfrom["maturity_over_six_months"] = df[base_filter & (
        df["Maturity"] == "Over 6 months"
)]

    return datesfrom

def xlsx_to_csv(filename):
    df = pd.read_excel(filename)
    return df.to_csv(index=False)

def prepare_report(df, currency_pair=("USD", "AUD")):
    data = {}

    instrument_types = [
        "Spot",
        "Outright forwards",
        "Currency swaps",
        "FX swaps",
        "Options",
        "Total (all instruments)"
    ]

    for instrument in instrument_types:
        data[instrument] = extract_fx_data(df, instrument=instrument, currency_pair=currency_pair)

    return data

def generate_report(dataset, currency="USD", year="2022"):
    wb = load_workbook("Template_Datathon.xlsx")
    sheet = wb['A2']
    errors = []
    
    currencies = ["AUD", "BRL", "CAD", "CHF", "CNY", "EUR", "GBP", "HKD", "INR", "JPY", "KRW", "MXN", "NOK", "NZD", "PLN", "RUB", "SEK", "SGD", "TRY", "TWD", "ZAR"]
    for index, currency2 in enumerate(currencies):
        report_column = prepare_report(dataset, (currency, currency2))
        
        # sheet.cell(9, 4 + index, report_column["Spot"]["total"][str(year)].sum())
        i = 10
        sheet.cell(i, 4 + index, report_column["Spot"]["reporting_dealers"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["reporting_dealers_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["reporting_dealers_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_non_reporting"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_inst_investor"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_hedge"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_official"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_others"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["other_fin_undistributed"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["non_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["non_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["non_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["total"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["broke_non_bank_ele"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["broke_other"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Spot"]["broke_retail"][str(year)].sum())
        i = i + 2
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["reporting_dealers"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["reporting_dealers_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["reporting_dealers_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_non_reporting"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_inst_investor"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_hedge"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_official"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_others"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["other_fin_undistributed"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["non_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["non_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["non_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["total"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["broke_non_bank_ele"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["broke_other"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["broke_retail"][str(year)].sum())
        i = i + 1
        # sheet.cell(i, 4 + index, report_column["Outright forwards"]["non_del_forwards"][str(year)].sum())
        i = i + 2
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["maturity_one_day"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["maturity_one_up_to_seven"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["maturity_seven_up_to_month"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["maturity_month_up_to_three"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["maturity_three_month_up_to_six"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Outright forwards"]["maturity_over_six_months"][str(year)].sum())
        i = i + 2
        sheet.cell(i, 4 + index, report_column["FX swaps"]["reporting_dealers"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["reporting_dealers_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["reporting_dealers_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_non_reporting"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_inst_investor"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_hedge"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_official"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_others"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["other_fin_undistributed"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["non_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["non_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["non_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["total"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["broke_non_bank_ele"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["broke_other"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["broke_retail"][str(year)].sum())
        i = i + 2
        sheet.cell(i, 4 + index, report_column["FX swaps"]["maturity_one_day"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["maturity_one_up_to_seven"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["maturity_seven_up_to_month"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["maturity_month_up_to_three"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["maturity_three_month_up_to_six"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["FX swaps"]["maturity_over_six_months"][str(year)].sum())

        i = i + 2
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["reporting_dealers"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["reporting_dealers_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["reporting_dealers_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_non_reporting"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_inst_investor"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_hedge"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_official"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_others"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["other_fin_undistributed"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["non_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["non_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["non_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["total"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["broke_non_bank_ele"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["broke_other"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Currency swaps"]["broke_retail"][str(year)].sum())

        i = i + 2
        sheet.cell(i, 4 + index, report_column["Options"]["reporting_dealers"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["reporting_dealers_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["reporting_dealers_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_non_reporting"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_inst_investor"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_hedge"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_official"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_others"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["other_fin_undistributed"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["non_fin"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["non_fin_local"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["non_fin_cross_border"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["total"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["broke_non_bank_ele"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["broke_other"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Options"]["broke_retail"][str(year)].sum())
        i = i + 2
        sheet.cell(i, 4 + index, report_column["Total (all instruments)"]["broke_non_bank_ele"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Total (all instruments)"]["broke_other"][str(year)].sum())
        i = i + 1
        sheet.cell(i, 4 + index, report_column["Total (all instruments)"]["broke_retail"][str(year)].sum())

    report_name = f"reports/Report-{year}-{currency}-{generate_uuid()}.xlsx"
    wb.save(report_name)
    return report_name, errors